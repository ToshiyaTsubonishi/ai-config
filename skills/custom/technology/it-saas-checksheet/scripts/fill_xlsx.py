#!/usr/bin/env python3
"""
fill_xlsx.py

SaaS・ASPチェックシート（template.xlsx）の 16行目以降（No.1〜13）の回答欄に、
JSON入力（answers/supplement）を転記して output.xlsx を生成する。

- Excel / LibreOffice 不要（openpyxlで.xlsxを直接編集）
- 既存の書式・結合セルは維持（結合セルの左上セルに値を書き込む）
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict

from openpyxl import load_workbook


SHEET_NAME = "SaaS・ASPチェックシート"

# No -> row mapping (16行目〜28行目)
ROW_BY_NO = {
    1: 16,
    2: 17,
    3: 18,
    4: 19,
    5: 20,
    6: 21,
    7: 22,
    8: 23,
    9: 24,
    10: 25,
    11: 26,
    12: 27,
    13: 28,
}

# Columns
COL_YESNO = "G"   # YES/NO
COL_DETAIL = "H"  # 回答内容（H〜K結合の左上）


def _normalize_yesno(value: Any) -> str:
    """
    Normalize yes/no variants to 'YES' or 'NO'.
    Accepts YES/NO (case-insensitive), True/False, 'Y'/'N', 'はい'/'いいえ'.
    """
    if value is None:
        raise ValueError("yes_no is missing")

    if isinstance(value, bool):
        return "YES" if value else "NO"

    s = str(value).strip()

    # Common variants
    if s.lower() in {"yes", "y", "true", "1", "はい", "有", "あり"}:
        return "YES"
    if s.lower() in {"no", "n", "false", "0", "いいえ", "無", "なし"}:
        return "NO"

    # Already normalized?
    if s in {"YES", "NO"}:
        return s

    raise ValueError(f"Invalid yes_no: {value!r} (must be YES/NO)")


def fill_checksheet(
    template_path: Path,
    output_path: Path,
    answers: Dict[str, Dict[str, Any]],
    supplement: str = "",
    sheet_name: str = SHEET_NAME,
) -> None:
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Fill answers
    for no_str, payload in answers.items():
        try:
            no = int(no_str)
        except Exception as e:
            raise ValueError(f"Answer key must be an integer string: {no_str!r}") from e

        if no not in ROW_BY_NO:
            # Ignore out-of-range keys to keep it robust for demos
            continue

        row = ROW_BY_NO[no]
        yesno_cell = f"{COL_YESNO}{row}"
        detail_cell = f"{COL_DETAIL}{row}"

        yes_no = _normalize_yesno(payload.get("yes_no"))
        detail = payload.get("detail", "")
        if detail is None:
            detail = ""

        ws[yesno_cell].value = yes_no
        ws[detail_cell].value = str(detail)

    # Supplement row (C29〜L29 merged; write to top-left C29)
    if supplement is None:
        supplement = ""
    ws["C29"].value = str(supplement)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _load_input_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Input JSON not found: {path}")
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, dict):
        raise ValueError("Input JSON must be an object")

    answers = data.get("answers")
    if not isinstance(answers, dict):
        raise ValueError("Input JSON must contain object field 'answers'")

    # Ensure payloads are dict
    for k, v in answers.items():
        if not isinstance(v, dict):
            raise ValueError(f"answers[{k!r}] must be an object")

    supplement = data.get("supplement", "")
    return {"answers": answers, "supplement": supplement}


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Fill SaaS・ASPチェックシート (rows 16-28) from JSON and output .xlsx"
    )
    p.add_argument(
        "--template",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "assets" / "template.xlsx",
        help="Path to template .xlsx (default: skills/saas-checksheet/assets/template.xlsx)",
    )
    p.add_argument(
        "--input",
        type=Path,
        required=True,
        help="Path to input JSON (must include 'answers' and optional 'supplement')",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=Path("output.xlsx"),
        help="Path to output .xlsx (default: ./output.xlsx)",
    )
    args = p.parse_args(argv)

    payload = _load_input_json(args.input)
    fill_checksheet(
        template_path=args.template,
        output_path=args.output,
        answers=payload["answers"],
        supplement=payload["supplement"],
    )
    print(f"OK: wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
